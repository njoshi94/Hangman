"""
Hangman Game Using SVR Machine Learning Methods 
    to predict Number of Tries needed to solve puzzle 
    Stores Data on DynamoDB using boto3 
    Includes custom GUI interface built with tkinter 
    Also can be played locally utilizing openpyxl library

"""

import tkinter as tk
import pandas as pd 
import numpy as np 
import random
import requests
import unidecode
import boto3 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.svm import SVR
from openpyxl.utils.dataframe import dataframe_to_rows
from boto3.dynamodb.conditions import Key

root = tk.Tk()

#Function to display guessed letters and _ for missing letters 
def disp_word(word):
    printer = ''
    for c in range(0,len(word)):
        if word['value'][c] == 0:
            printer = printer + ' _ '
        else:
            printer = printer + ' ' + word['letter'][c] + ' '
    Printer.set(printer)


#Code for actual game 
def game(block):
    
    #Initializing AWS for later 
    #Resource used over client for flexibility
    DB = boto3.resource('dynamodb')
    
    #Hashtable of all single character entries 
    hashtable = {}
    for i in range(33,127):
        hashtable[chr(i)]= 0
    hashtable[' '] = 0
    
    if OnLine.get() == 1:
        
        #Pulling necessary tables from AWS 
        TableName = "Tries"

        table = DB.Table(TableName)
        response = table.scan()
        
        #Converting to DataFrame and arranging columns appropriately
        df = response['Items']
        df = pd.DataFrame(df)        
        cols = ['ID', 'Num_Letters', 'Num_Tries']
        df = df[cols]        

        TableName = "Fails"

        table = DB.Table(TableName)
        response = table.scan()
        
        df1 = response['Items']
        df1 = pd.DataFrame(df1)
        cols = ['Letter', 'Fails']
        df1 = df1[cols]

        TableName = "Guesses"

        table = DB.Table(TableName)
        response = table.scan()
        
        df2 = response['Items']
        df2 = pd.DataFrame(df2)
        cols = ['Letter', 'Guesses'] 
   
    
    if OnLine.get() == 0:
        filename = 'ML1.xlsx'
        
        #Checking if Excel Document with data exists
        #Creates document if it does not exist 
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
        except FileNotFoundError:
            headers_row = ['Num_Letters', 'Num_Tries']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)
            wb.create_sheet()
            ws1 = wb.worksheets[1]
            ws1.append(['Letter', 'Fails'])
            for i in range(33,127):
                ws1.append([chr(i), 0])
            ws1.append([' ',0])
            ws1.append(['Total', 0])
            wb.create_sheet()
            ws2 = wb.worksheets[2]
            ws2.append(['Letter', 'Guesses'])
            for i in range(33,127):
                ws2.append([chr(i), 0])
            ws2.append([' ', 0])
            ws2.append(['Total', 0])                
            wb.save(filename)
            
        df = pd.read_excel(filename)
        df1 = pd.read_excel(filename, sheet_name = 'Sheet1')
        df2 = pd.read_excel(filename, sheet_name = 'Sheet2')
        
    
    """ main function """
    #preparing for game
    if HardMode.get() == 0:
        if OnLine == 0:
            #Word is randomly selected from list of words            
            Word_List = pd.read_excel('Words.xlsx')
            r = random.randint(0,len(Word_List))
            temp = Word_List['Words'][r]
        else:
            #Word is pulled from Dynamodb table 
            r = random.randint(1,500)
            table = DB.Table("Words")
            response = table.query(
                KeyConditionExpression = Key('ID').eq(r)
                    )
            data = response["Items"]
            temp = data[0]['Word']
    else:
        #If Hardmode, pulls word from website of difficult words 
        URL = 'https://randomword.com/'
        page = requests.get(URL)        
        soup = BeautifulSoup(page.content, 'html.parser')
        word = soup.find(id = 'random_word').get_text()        
        temp = unidecode.unidecode(word)
    
    if OnLine == 0:
        #If Local and less than 25 games have been played
        #calculate straight average number of tries 
        #for words of length X
        #If no records exist, set number of tries to 12 
        if len(df) < 25:
            if len(df[df['Num_Letters'] == len(temp)]['Num_Tries']) > 0:
                p = round(np.mean(df[df['Num_Letters'] == len(temp)]['Num_Tries'].values))
            else:
                Special.set("No Previous Data")
                p = 12 
        else:
            #If more than 25 games played, use SVR to calculate number of tries
            #Expected graph is nonlinear / polynomic and SVR accounts for this
            #Also can handle large spread of points at each x-value 
            X = df.iloc[:, 0:1].values
            y = df.iloc[:,1].values        
            
            regressor = SVR(gamma = 'scale')
            regressor.fit(X,y)
            
            array = np.array([len(temp)])
            pr = regressor.predict(array.reshape(1,-1))
            
            p = pr[0]
            
            if max(df1['Fails']) > 25:
                #If more than 25 games lost, will increase number of tries
                #based on if word has problematic letters 
                Top10Fails = df1.nlargest(11, 'Fails')
                FD = max(df1['Fails'])
                
                for c in range (0, len(temp)):
                    if temp[c] in Top10Fails['Letter'].values:
                        #p = p + 
                        FailAddition = Top10Fails.loc[Top10Fails['Letter'] == temp[c], 'Fails'] / FD
                        FTV = FailAddition.values
                        p = p + FTV[0]
                        
            if max(df2['Guesses']) > 100:
                #If more than 100 games played, will decrease number of tries
                #based on if word has common letters 
                Top10Guesses = df2.nlargest(11, 'Guesses')
                GD = max(df2['Guesses']) * 10 
                #Can we add ML to determine that?
                for c in range (0, len(temp)):
                    if temp[c] in Top10Guesses['Letter'].values:
                        #p = p + 
                        GuessSubtraction = Top10Guesses.loc[Top10Guesses['Letter'] == temp[c], 'Guesses'] / GD
                        GTV = GuessSubtraction.values
                        p = p - GTV[0]
    else:
        #Same methodology as above for SVR and beyond 
        # columns have to be set dynamically because DynamoDB table
        #has extra column and playing in local mode will crash 
        if len(df.columns) == 1:
            width = 1 
        else: 
            width = len(df.columns) - 1
        X = df.iloc[:, width - 1:width].values
        y = df.iloc[:,width].values        
        
        regressor = SVR(gamma = 'scale')
        regressor.fit(X,y)
        
        array = np.array([len(temp)])
        pr = regressor.predict(array.reshape(1,-1))
        
        p = pr[0]
        
        if max(df1['Fails']) > 25:
            Top10Fails = df1.nlargest(11, 'Fails')
            FD = max(df1['Fails'])
            
            for c in range (0, len(temp)):
                if temp[c] in Top10Fails['Letter'].values:
                    #p = p + 
                    FailAddition = Top10Fails.loc[Top10Fails['Letter'] == temp[c], 'Fails'] / FD
                    q = FailAddition.idxmax()
                    p = p + FailAddition[q]         
        if max(df2['Guesses']) > 100:
            Top10Guesses = df2.nlargest(11, 'Guesses')
            GD = max(df2['Guesses']) * 10 
            #Can we add ML to determine that?
            for c in range (0, len(temp)):
                if temp[c] in Top10Guesses['Letter'].values:
                    #p = p + 
                    GuessSubtraction = Top10Guesses.loc[Top10Guesses['Letter'] == temp[c], 'Guesses'] / GD
                    q = GuessSubtraction.idxmax()
                    p = p - GuessSubtraction[q]                
    p = round(p)
    
    #Create Word data structure that contains the letters and key if it has 
    #already been guessed
    word = np.empty(shape = (len(temp),2), dtype = str)
    for c in range(0, len(temp)):
        word[c] = (temp[c],0)
    word = pd.DataFrame(word, columns = ['letter', 'value'])
    word['value'] = pd.to_numeric(word['value'])
    
    i = 0
    while i < p:
        #Main Loop 
        j = 1
        Tries.set("Number of Tries Remaining: " + str(int(p- i)))
        disp_word(word)
        Command.set("Please input a letter: ")
        root.wait_variable(block)
        letter = LetterEntry.get()
        LetterEntry.delete(0, 'end')
        block.set('True')
        Special.set("")
        while len(letter) > 1 and letter != 'quit':
            if letter =='list':
                #Creates a list of letters already guessed
                Lister = ''
                for key, value in hashtable.items():
                    if value == 1:
                        Lister = Lister + key + ', '
                Lister = Lister[:-2]                        
                Special.set(Lister)
            Command.set("Please input a letter: ")
            root.wait_variable(block)
            letter = LetterEntry.get()
            LetterEntry.delete(0, 'end')
            block.set('True')
            Special.set("")
        if letter == 'quit':
            break
        if len(letter) == 1 and letter != ' ':
            if hashtable[letter] == 1:
                #Checks if letter has already been set 
                Special.set("You already guessed that!")
                j = 0
            else:
                hashtable.update({letter: 1})
                for c in range(0,len(word)):
                    if letter == word.loc[c, ('letter')]:
                        word.loc[c, ('value')] = 1
                        j = 0 
            
        if min(word['value']) == 1:
            #If all letters have been guessed, break loop 
            break
        #Only increment counter if new letter not in word 
        i = i + j 
    
    Printer.set("")
    Tries.set("")
    Command.set("")
    Special.set("")
    
    
    
    if i < p and letter != 'quit':
        Printer.set("Congrats! You won!")
        Tries.set("The word was: " + temp)
       
        if OnLine == 0:
            #Updates local tables with results of winning game 
            ws.append([len(temp), i])    
            
            for key, value in hashtable.items():
                if value == 1:
                    df2.loc[df2['Letter'] == key, 'Guesses'] +=1
            df2.loc[df2['Letter'] == 'Total', 'Guesses'] +=1 
            delt = wb['Sheet2']
            wb.remove(delt)
            ws2 = wb.create_sheet('Sheet2')
            for r in dataframe_to_rows(df2, index=False, header=True):
                ws2.append(r)
            wb.save(filename)
        else:
            #Updates DynamoDB tables with results of winning game 
            TableName = "Tries"
            table = DB.Table(TableName)
            Val = df.columns
            response = table.put_item(
                Item = {
                    'ID': max(df[Val[0]]) + 1,
                    'Num_Letters' : len(temp),
                    'Num_Tries': i 
                    }            
                )
            
            TableName = "Guesses"
            table = DB.Table(TableName)
            for key, value in hashtable.items():
                if value == 1:
                    df2.loc[df2['Letter'] == key, 'Guesses'] +=1
            df2.loc[df2['Letter'] == 'Total', 'Guesses'] +=1
            i = 0 
            while i < len(df2):
                response = table.update_item(
                    Key = {
                        'Letter' : df2['Letter'][i] 
                        },
                    UpdateExpression = 'SET Guesses =  :val1',
                    ExpressionAttributeValues = {
                        ':val1' : df2['Guesses'][i]
                        }
                    )
                i = i + 1 

        Command.set("Enter 'y' to play agan: ")
        root.wait_variable(block)
        letter = LetterEntry.get()
        LetterEntry.delete(0, 'end')
        block.set('True')
        
    elif letter == 'quit':
        pass
    else:
        Printer.set("Sorry! The correct word was: ")
        Tries.set(temp)
        if OnLine == 0:
            #Updates local tables with results of losing game 
            ws.append([len(temp), 12])
            k = 0 
            for c in range(0,len(word)):
                if word['value'][c] == 0:
                    m = word['letter'][c]            
                    df1.loc[df1['Letter'] == m , 'Fails'] +=1
                    k = k + 1 
            df1.loc[df1['Letter'] == 'Total', 'Fails'] += k
            delt = wb['Sheet1']
            wb.remove(delt)
            ws1 = wb.create_sheet('Sheet1')
            for r in dataframe_to_rows(df1, index=False, header=True):
                ws1.append(r)
            for key, value in hashtable.items():
                if value == 1:
                    df2.loc[df2['Letter'] == key, 'Guesses'] +=1
            df2.loc[df2['Letter'] == 'Total', 'Guesses'] +=1 
            delt = wb['Sheet2']
            wb.remove(delt)
            ws2 = wb.create_sheet('Sheet2')
            for r in dataframe_to_rows(df2, index=False, header=True):
                ws2.append(r)
            wb.save(filename)  
        else:
            #Updates DynamoDB tables with results of losing game 
            TableName = "Tries"
            Val = df.columns
            table = DB.Table(TableName)
            response = table.put_item(
                Item = {
                    'ID': max(df[Val[0]]) + 1,
                    'Num_Letters' : len(temp),
                    'Num_Tries': 12 
                    }            
                )
            
            TableName = "Fails"
            table = DB.Table(TableName)
            k = 0 
            for c in range(0,len(word)):
                if word['value'][c] == 0:
                    m = word['letter'][c]            
                    df1.loc[df1['Letter'] == m , 'Fails'] +=1
                    k = k + 1 
            df1.loc[df1['Letter'] == 'Total', 'Fails'] += k
            i = 0
            while i < len(df1):
                response = table.update_item(
                    Key = {
                        'Letter' : df1['Letter'][i] 
                        },
                    UpdateExpression = 'SET Fails = :val1',
                    ExpressionAttributeValues = {
                        ':val1' : df1['Fails'][i]
                        }
                    )
                i = i + 1
            
            TableName = "Guesses"
            table = DB.Table(TableName)
            for key, value in hashtable.items():
                if value == 1:
                    df2.loc[df2['Letter'] == key, 'Guesses'] +=1
            df2.loc[df2['Letter'] == 'Total', 'Guesses'] +=1 
            i = 0
            while i < len(df2):
                response = table.update_item(
                    Key = {
                        'Letter' : df2['Letter'][i] 
                        },
                    UpdateExpression = 'SET Guesses = :val1',
                    ExpressionAttributeValues = {
                        ':val1' : df2['Guesses'][i]
                        }
                    )
                i = i + 1 
        
        Command.set("Enter 'y' to play agan: ")
        root.wait_variable(block)
        letter = LetterEntry.get()
        LetterEntry.delete(0, 'end')
        block.set('True')
    
    
    Printer.set("")
    Tries.set("")
    Command.set("")
    Special.set("")    
    if letter == 'y':
        game(block)

def gamestart():
    #Disables Options Buttons and creates screen for game 
    StartGame.config(state = 'disabled')
    Checker.config(state = 'disabled')
    Checker2.config(state = 'disabled')
    LetterEntry.config(state = 'normal')
    InnerFrame.config(highlightthickness = 1)
    InnerFrame.config(bg = '#9f9fad')
    WordDisplay.config(bg = '#9f9fad')
    TriesDisplay.config(bg = '#9f9fad')
    CommandDisplay.config(bg = '#9f9fad')
    SpecialDisplay.config(bg = '#9f9fad')
    game(block)
    #Returns screen back to inital state 
    StartGame.config(state = 'active')
    Checker.config(state = 'active')
    Checker2.config(state = 'active')
    LetterEntry.config(state = 'disabled')
    InnerFrame.config(highlightthickness = 0)
    InnerFrame.config(bg = '#81e4da')
    WordDisplay.config(bg = '#81e4da')
    TriesDisplay.config(bg = '#81e4da')
    CommandDisplay.config(bg = '#81e4da')
    SpecialDisplay.config(bg = '#81e4da')

def InputLetter():
    block.set(False)

def enableHM():
    if Checker["text"] == 'Hard Mode: Disabled':
        HardMode.set(1)
        Checker.config(text = 'Hard Mode: Enabled')
    else:
        HardMode.set(0)
        Checker.config(text = 'Hard Mode: Disabled') 

def enableOF():
    if Checker2["text"] == 'Online Mode':
        OnLine.set(0)
        Checker2.config(text = 'Offline Mode')
    else:
        OnLine.set(1)
        Checker2.config(text = 'Online Mode')
        
#All tkinter components 
block = tk.BooleanVar(root, True)

HardMode = tk.IntVar(0)

OnLine = tk.IntVar()
OnLine.set(1)

Printer = tk.StringVar()
Printer.set("")

Tries = tk.StringVar()
Tries.set("")

Command = tk.StringVar()
Command.set("")

Special = tk.StringVar()
Special.set("")

root.title("Hangman Game")

canvas = tk.Canvas(root, height = 500, width = 500, bg = '#47e5bc').pack()

Frame = tk.Frame(root, bg = '#81e4da', highlightbackground = 'black', highlightthickness = 1)
Frame.place(relx = .1, rely = .1, relheight = .8, relwidth = .8)

InnerFrame = tk.Frame(root, bg = '#81e4da', highlightbackground = 'black', highlightthickness = 0)
InnerFrame.place(anchor = 'center', relx = .5, rely = .325, relheight = .25, relwidth = .5)

StartGame = tk.Button(root, text = 'Start Game', bg = '#aecfdf', command = gamestart)
StartGame.place(anchor = 'center', rely = .7, relx = .5)

Checker = tk.Button(root, text = 'Hard Mode: Disabled', bg = '#aecfdf', command = enableHM)
Checker.place(anchor = 'center', rely = .77, relx = .5)

Checker2 = tk.Button(root, text = 'Online Mode', bg = '#aecfdf', command = enableOF)
Checker2.place(anchor = 'center', rely = .84, relx = .5)

LetterEntry = tk.Entry(root, state = 'disabled')
LetterEntry.bind('<Return>', (lambda event: InputLetter()))
LetterEntry.place(anchor = 'center', relx = .5, rely = .5)

WordDisplay = tk.Message(root, textvariable = Printer, width = 30000, bg = '#81e4da')
WordDisplay.place(anchor = 'center', relx = .5, rely = .25)

TriesDisplay = tk.Message(root, textvariable = Tries, width = 30000, bg = '#81e4da')
TriesDisplay.place(anchor = 'center', relx = .5, rely = .3)

CommandDisplay = tk.Message(root, textvariable = Command, width = 30000, bg = '#81e4da')
CommandDisplay.place(anchor = 'center', relx = .5, rely = .35)

SpecialDisplay = tk.Message(root, textvariable = Special, width = 30000, bg = '#81e4da')
SpecialDisplay.place(anchor = 'center', relx = .5, rely = .4)

root.mainloop()